"""
One-time data import script from Tacoma_Service_20251125.xlsx.

Usage:
    python scripts/import_spreadsheet.py path/to/Tacoma_Service_20251125.xlsx

Idempotent — safe to run multiple times (uses upsert logic via checking existing records).
"""
import sys
import uuid
import logging
from datetime import date, datetime, timedelta
from decimal import Decimal
from pathlib import Path

import openpyxl
from sqlalchemy import create_engine, select
from sqlalchemy.orm import Session

# Add parent dir to path so we can import app modules
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from app.database import Base
from app.models import Vehicle, OilChange, ServiceRecord, IntervalItem, Observation, AppSettings
from app.models.interval_item import IntervalItemType
from app.config import settings

logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
log = logging.getLogger(__name__)

# Excel serial date epoch (Jan 0, 1900 — Excel treats 1900-01-01 as serial 1)
EXCEL_EPOCH = datetime(1899, 12, 30)


def get_sync_engine():
    return create_engine(settings.DATABASE_URL_SYNC)


def safe_int(value) -> int | None:
    """Convert a value to int, handling 'Unknown', None, and float values."""
    if value is None:
        return None
    if isinstance(value, str):
        cleaned = value.strip().replace(",", "")
        if cleaned.lower() in ("unknown", "n/a", ""):
            return None
        try:
            return int(float(cleaned))
        except ValueError:
            return None
    try:
        return int(value)
    except (ValueError, TypeError):
        return None


def safe_date(value) -> date | None:
    """Convert a value to date, handling Excel serial numbers."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        # Excel serial date number
        try:
            serial = int(value)
            if 1 < serial < 200000:  # reasonable date range
                return (EXCEL_EPOCH + timedelta(days=serial)).date()
        except (ValueError, OverflowError):
            pass
    if isinstance(value, str):
        for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%m/%d/%y"):
            try:
                return datetime.strptime(value.strip(), fmt).date()
            except ValueError:
                continue
    return None


def safe_float(value) -> float | None:
    """Convert a value to float."""
    if value is None:
        return None
    if isinstance(value, str):
        cleaned = value.strip().replace("$", "").replace(",", "")
        if cleaned in ("", "-"):
            return None
        try:
            return float(cleaned)
        except ValueError:
            return None
    try:
        return float(value)
    except (ValueError, TypeError):
        return None


def import_vehicles(ws, session: Session) -> dict[str, uuid.UUID]:
    """Import from 'Truck Details' sheet (key-value layout: label in col A, value in col B)."""
    log.info("Importing vehicles...")
    vehicle_map = {}

    # Read key-value pairs from the sheet
    kv = {}
    for row in ws.iter_rows(min_row=1, values_only=True):
        if row and row[0] and len(row) > 1:
            key = str(row[0]).strip().lower()
            kv[key] = row[1]

    year = safe_int(kv.get("year"))
    make = str(kv.get("make", "")).strip()
    model = str(kv.get("model", "")).strip()
    trim = str(kv.get("trim", "")).strip() or None
    color = str(kv.get("color", "")).strip() or None
    vin = str(kv.get("vin", "")).strip() or None
    mileage = safe_int(kv.get("current mileage") or kv.get("mileage")) or 0

    if not year or not make:
        log.warning(f"  Skipping vehicle: year={year}, make={make}")
        return vehicle_map

    # Check if vehicle already exists by VIN or year+make+model
    existing = None
    if vin:
        existing = session.execute(select(Vehicle).where(Vehicle.vin == vin)).scalar_one_or_none()
    if not existing:
        existing = session.execute(
            select(Vehicle).where(Vehicle.year == year, Vehicle.make == make, Vehicle.model == model)
        ).scalar_one_or_none()

    if existing:
        log.info(f"  Vehicle already exists: {year} {make} {model}")
        vehicle_map[f"{year} {make} {model}"] = existing.id
        if mileage and mileage > existing.current_mileage:
            existing.current_mileage = mileage
    else:
        vehicle = Vehicle(
            year=year, make=make, model=model, trim=trim,
            color=color, vin=vin, current_mileage=mileage,
        )
        session.add(vehicle)
        session.flush()
        vehicle_map[f"{year} {make} {model}"] = vehicle.id
        log.info(f"  Imported: {year} {make} {model}")

    return vehicle_map


def find_header_row(ws, expected_first_col: str) -> int:
    """Find the row number containing the header by looking for expected_first_col in column A."""
    for i, row in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
        if row and row[0] and str(row[0]).strip().lower() == expected_first_col.lower():
            return i
    return 1  # fallback


def import_oil_changes(ws, session: Session, vehicle_id: uuid.UUID):
    """Import from 'Oil Changes' sheet. Header row starts with 'Date'."""
    log.info("Importing oil changes...")
    imported, skipped = 0, 0

    header_row = find_header_row(ws, "Date")
    rows = list(ws.iter_rows(min_row=header_row + 1, values_only=True))
    for row in rows:
        if not row or not row[0]:
            continue

        service_date = safe_date(row[0])
        if not service_date:
            skipped += 1
            continue

        facility = str(row[1]).strip() if len(row) > 1 and row[1] else None
        odometer = safe_int(row[2]) if len(row) > 2 else None
        if odometer is None:
            skipped += 1
            log.warning(f"  Skipping oil change on {service_date}: no odometer")
            continue

        interval_miles = safe_int(row[3]) if len(row) > 3 else None
        interval_months = safe_float(row[4]) if len(row) > 4 else None
        notes = str(row[5]).strip() if len(row) > 5 and row[5] else None

        # Check for duplicate
        existing = session.execute(
            select(OilChange).where(
                OilChange.vehicle_id == vehicle_id,
                OilChange.service_date == service_date,
                OilChange.odometer == odometer,
            )
        ).scalar_one_or_none()

        if existing:
            skipped += 1
            continue

        oc = OilChange(
            vehicle_id=vehicle_id, service_date=service_date, facility=facility,
            odometer=odometer, interval_miles=interval_miles,
            interval_months=interval_months, notes=notes,
        )
        session.add(oc)
        imported += 1

    log.info(f"  Oil changes: {imported} imported, {skipped} skipped")


def import_service_records(ws, session: Session, vehicle_id: uuid.UUID):
    """Import from 'Other Services' sheet. Header row starts with 'Date'."""
    log.info("Importing service records...")
    imported, skipped = 0, 0

    header_row = find_header_row(ws, "Date")
    rows = list(ws.iter_rows(min_row=header_row + 1, values_only=True))
    for row in rows:
        if not row or not row[0]:
            continue

        service_date = safe_date(row[0])
        if not service_date:
            skipped += 1
            continue

        facility = str(row[1]).strip() if len(row) > 1 and row[1] else None
        odometer = safe_int(row[2]) if len(row) > 2 else None

        # Services performed — may be in one cell (newline-separated) or multiple columns
        services = []
        if len(row) > 3 and row[3]:
            raw = str(row[3])
            services = [s.strip() for s in raw.replace("\r\n", "\n").split("\n") if s.strip()]

        notes = str(row[4]).strip() if len(row) > 4 and row[4] else None

        # Check for duplicate
        existing = session.execute(
            select(ServiceRecord).where(
                ServiceRecord.vehicle_id == vehicle_id,
                ServiceRecord.service_date == service_date,
                ServiceRecord.facility == facility,
            )
        ).scalar_one_or_none()

        if existing:
            skipped += 1
            continue

        record = ServiceRecord(
            vehicle_id=vehicle_id, service_date=service_date, facility=facility,
            odometer=odometer, services_performed=services or None, notes=notes,
        )
        session.add(record)
        imported += 1

    log.info(f"  Service records: {imported} imported, {skipped} skipped")


def import_interval_items(ws, session: Session, vehicle_id: uuid.UUID):
    """Import from 'Interval Tracking' sheet.

    Spreadsheet layout:
      Rows 1-7: Summary stats (Current Mileage, etc.)
      Row 8 (approx): "Scheduled" label
      Row 9: Header — Item, Type, Last Service Date, Last Service Miles,
              Next Service Mileage, Recommended Interval, Elapsed Miles, Diff,
              Approx. Cost, Next Service, Months Elapsed, Years Elapsed
      Row 10+: Data rows
    """
    log.info("Importing interval items...")
    imported, skipped = 0, 0

    header_row = find_header_row(ws, "Item")
    rows = list(ws.iter_rows(min_row=header_row + 1, values_only=True))
    for row in rows:
        if not row or not row[0]:
            continue

        name = str(row[0]).strip()
        if not name or name.lower() in ("sub total", "total"):
            continue

        # Col B: Type (Regular / Ad-Hoc)
        item_type_str = str(row[1]).strip().lower() if len(row) > 1 and row[1] else "regular"
        item_type = IntervalItemType.AD_HOC if "ad" in item_type_str else IntervalItemType.REGULAR

        # Col C: Last Service Date (Excel serial)
        last_date = safe_date(row[2]) if len(row) > 2 else None
        # Col D: Last Service Miles
        last_miles = safe_int(row[3]) if len(row) > 3 else None
        # Col E: Next Service Mileage
        next_miles = safe_int(row[4]) if len(row) > 4 else None
        # Col F: Recommended Interval
        interval_miles = safe_int(row[5]) if len(row) > 5 else None
        # Col I: Approx. Cost
        cost = safe_float(row[8]) if len(row) > 8 else None

        # Check for duplicate
        existing = session.execute(
            select(IntervalItem).where(
                IntervalItem.vehicle_id == vehicle_id,
                IntervalItem.name == name,
            )
        ).scalar_one_or_none()

        if existing:
            skipped += 1
            continue

        item = IntervalItem(
            vehicle_id=vehicle_id, name=name, type=item_type,
            last_service_date=last_date, last_service_miles=last_miles,
            recommended_interval_miles=interval_miles, next_service_miles=next_miles,
            due_soon_threshold_miles=500, estimated_cost=cost,
        )
        session.add(item)
        imported += 1

    log.info(f"  Interval items: {imported} imported, {skipped} skipped")


def import_observations(ws, session: Session, vehicle_id: uuid.UUID):
    """Import from 'Misc - Observations' sheet. Header row starts with 'Date'."""
    log.info("Importing observations...")
    imported, skipped = 0, 0

    header_row = find_header_row(ws, "Date")
    rows = list(ws.iter_rows(min_row=header_row + 1, values_only=True))
    for row in rows:
        if not row or not row[0]:
            continue

        obs_date = safe_date(row[0])
        if not obs_date:
            skipped += 1
            continue

        odometer = safe_int(row[1]) if len(row) > 1 else None
        observation = str(row[2]).strip() if len(row) > 2 and row[2] else None
        if not observation:
            skipped += 1
            continue

        resolved = False
        if len(row) > 3 and row[3]:
            resolved = str(row[3]).strip().lower() in ("yes", "true", "resolved", "1")

        resolved_date = safe_date(row[4]) if len(row) > 4 else None

        # Check for duplicate
        existing = session.execute(
            select(Observation).where(
                Observation.vehicle_id == vehicle_id,
                Observation.observation_date == obs_date,
                Observation.observation == observation,
            )
        ).scalar_one_or_none()

        if existing:
            skipped += 1
            continue

        obs = Observation(
            vehicle_id=vehicle_id, observation_date=obs_date, odometer=odometer,
            observation=observation, resolved=resolved, resolved_date=resolved_date,
        )
        session.add(obs)
        imported += 1

    log.info(f"  Observations: {imported} imported, {skipped} skipped")


# Mapping of expected sheet names to import functions
SHEET_MAP = {
    "Oil Changes": import_oil_changes,
    "Other Services": import_service_records,
    "Interval Tracking": import_interval_items,
    "Misc - Observations": import_observations,
}


def find_sheet(wb, candidates: list[str]):
    """Find a sheet by name, supporting partial matches."""
    for name in wb.sheetnames:
        for candidate in candidates:
            if name == candidate or name.startswith(candidate):
                return wb[name], name
    return None, None


def main(xlsx_path: str):
    log.info(f"Opening {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    log.info(f"Sheets found: {wb.sheetnames}")

    engine = get_sync_engine()

    with Session(engine) as session:
        # Import vehicles first
        vehicle_sheet, sheet_name = find_sheet(wb, ["Truck Details", "Vehicle Details", "Vehicle"])

        if not vehicle_sheet:
            log.error("No vehicle details sheet found. Expected 'Truck Details'.")
            sys.exit(1)

        vehicle_map = import_vehicles(vehicle_sheet, session)
        if not vehicle_map:
            log.error("No vehicles imported.")
            sys.exit(1)

        # Use the first (and likely only) vehicle
        vehicle_id = list(vehicle_map.values())[0]
        log.info(f"Using vehicle ID: {vehicle_id}")

        # Import each data sheet (with partial name matching)
        for sheet_key, import_fn in SHEET_MAP.items():
            ws, matched_name = find_sheet(wb, [sheet_key])
            if ws:
                log.info(f"Found sheet '{matched_name}' for '{sheet_key}'")
                import_fn(ws, session, vehicle_id)
            else:
                log.warning(f"Sheet '{sheet_key}' not found, skipping")

        session.commit()
        log.info("Import complete.")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python scripts/import_spreadsheet.py <path-to-xlsx>")
        sys.exit(1)
    main(sys.argv[1])
